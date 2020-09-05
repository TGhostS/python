import telebot
import openpyxl
import json
from openpyxl import load_workbook
from telebot import types
import datetime
import keyboard
from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton
token ="1327620351:AAFuPqqT81BgTp0vaw7dzP1UvgENEUltJHo"
bot = telebot.TeleBot(token)
global flag
flag = False
data = {
    'chat.id':0,
    'Type_clinic':'Type',
    'FIO_doctor':'FIO',
    
}
#подключаем 2 входные кнопки
@bot.message_handler(commands=['start'])
def start_message(message):
    if flag == False:
        keyboard = telebot.types.ReplyKeyboardMarkup(True)
        keyboard.row('Зарегистрироваться', 'Войти')
        bot.send_message(message.chat.id, 'Здравствуйте! Я - Бот, могу помочь Вам записаться к врачу', reply_markup=keyboard)


@bot.message_handler(content_types=['text'])
def run(message):
    if message.text == 'Войти':
        msg = bot.send_message(message.chat.id,'Введите номер телефона и пароль (Образец: +XXXXXXXXXXX *********)')
        bot.register_next_step_handler(msg, switch)
    if message.text == 'Зарегистрироваться':
        msg = bot.send_message(message.chat.id,'чтобы зарегистрироваться надо ввести номер полиса, номер телефона, пароль, фамилия(Образец: XXXXXXXXXXX +XXXXXXXXXXX ********* Иванов')
        bot.register_next_step_handler(msg,proverka)
    if message.text == 'Выйти':
        keyboard = telebot.types.ReplyKeyboardMarkup(True)
        keyboard.row('Зарегистрироваться', 'Войти')
        bot.send_message(message.chat.id, 'Здравствуйте! Я - Бот, могу помочь Вам записаться к врачу', reply_markup=keyboard)

    
def switch(message):
    razres = message.text.split()
    reg = load_workbook('./Register.xlsx')
    sheet = reg.get_sheet_by_name('Лист1')
    i = 2
    phone = razres[0]
    password = razres[1]
    while True:
        global flag
        if sheet['A'+str(i)].value == password and sheet['D'+str(i)].value == phone:
            keyboard1 = telebot.types.ReplyKeyboardMarkup(True)
            keyboard1.row('Выйти')
            bot.send_message(message.chat.id,"Вы успешно вошли",reply_markup=keyboard1)
            key = types.InlineKeyboardMarkup()
            but_4 = types.InlineKeyboardButton(text="Государственная больница", callback_data="Государственная больница")
            but_5= types.InlineKeyboardButton(text="Частная больница", callback_data="Частная больница")
            key.add(but_4, but_5)
            bot.send_message(message.chat.id, "Выберите тип больницы", reply_markup=key)
            flag = True
            break
        if sheet['A'+str(i)].value == None:
            bot.send_message(message.chat.id,"Неправильный логин или пароль")
            break
        i += 1
 

@bot.callback_query_handler(func=lambda c:True)
def inlin(d):
    if d.data == 'Государственная больница':
        bot.delete_message(d.message.chat.id, d.message.message_id)
        key = types.InlineKeyboardMarkup()
        but_1 = types.InlineKeyboardButton(text="Запись к врачу", callback_data="Запись к врачу")
#TODO Реализовать сохранение информации о том , в какую болницу идет запись
        bot.send_message(d.message.chat.id, "Выберите кнопку", reply_markup = key)
        but_2 = types.InlineKeyboardButton(text="Получение направления", callback_data="Получение направления")
        but_3 = types.InlineKeyboardButton(text="Просмотр рейтинга и отзывов врачей", callback_data="Просмотр рейтинга и отзывов врачей")
        key.add(but_1, but_2, but_3)
        bot.send_message(d.message.chat.id, "Вы выбрали тип - государственная больница", reply_markup=key)
    if d.data == 'Частная больница':
        bot.delete_message(d.message.chat.id, d.message.message_id)
        key = types.InlineKeyboardMarkup()
        but_1 = types.InlineKeyboardButton(text="Запись к врачу", callback_data="Запись к врачу")
        but_2 = types.InlineKeyboardButton(text="Получение направления", callback_data="Получение направления")
        but_3 = types.InlineKeyboardButton(text="Просмотр рейтинга и отзывов врачей", callback_data="Просмотр рейтинга и отзывов врачей")
        key.add(but_1, but_2, but_3)
        bot.send_message(d.message.chat.id, "Вы выбрали тип - частная больница", reply_markup=key)
    if d.data == 'Запись к врачу':
        bot.delete_message(d.message.chat.id, d.message.message_id)
        createButtons("Психолог Педиатр",d,"Психолог Педиатр")
##кнопки, сменяющие запись к врачу
#@bot.callback_query_handler(func=lambda c:True)
#def inline(s):
#    if s.data == 'Запись к врачу':
#        bot.delete_message(s.message.chat.id, s.message.message_id)
#        key = types.InlineKeyboardMarkup()
#        but_6 = types.InlineKeyboardButton(text = "Терапевт", callback_data = "Терапевт")
#        but_7 = types.InlineKeyboardButton(text="Психолог", callback_data = "Психолог") 
#        key.add(but_6, but_7)
#        bot.send_message(s.message.chat.id, "Выберите кнопку", reply_markup = key)


def proverka(message):
    reg = load_workbook('./Register.xlsx')
    sheet = reg.get_sheet_by_name('Лист1')
    text = message.text.split()
    polise = text[0]
    password = text[2]
    phone = text[1]
    surname = text[3]
    i = 2  
    while True:
        if sheet['A'+str(i)].value == None:
            global flag
            sheet['A'+str(i)].value = password
            sheet['B'+str(i)].value = polise
            sheet['C'+str(i)].value = surname
            sheet['D'+str(i)].value = phone
            reg.save('./Register.xlsx')
            bot.send_message(message.chat.id,'Вы успешно зарегистрировались')
            key = types.InlineKeyboardMarkup()
            but_1 = types.InlineKeyboardButton(text="Запись к врачу", callback_data="Запись к врачу")
            but_2 = types.InlineKeyboardButton(text="Получение направления", callback_data="Получение направления")
            but_3 = types.InlineKeyboardButton(text="Просмотр рейтинга и отзывов врачей", callback_data="Просмотр рейтинга и отзывов врачей")
            key.add(but_1, but_2, but_3)
            bot.delete_message(message.chat.id, message.message_id)
            flag = True
            bot.send_message(message.chat.id, "Выберите кнопку", reply_markup=key)

            break
        i += 1

def createButtons(Name,message,callback):
    key = types.InlineKeyboardMarkup()
    counts = len(Name.split())
    names = Name.split()
    call = callback.split()
    i = 0
    while True:
            if i == counts:
                bot.send_message(message.chat.id,"Выбери кнопку",reply_markup=key)
                break
            but = types.InlineKeyboardButton(text=names[i],callback_data=call[i])
            key.add(but)
            


bot.polling(timeout=5,interval=0)